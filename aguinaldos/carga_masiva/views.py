from django.contrib import messages
from django.shortcuts import render, redirect
from .forms import CargaMasivaForm
from .models import Producto
import openpyxl
from .validators import validar_precio, validar_cantidad  # Asegúrate de importar ambos validadores
from django.conf import settings
import os
from django.core.files.storage import default_storage  # Para manejar el almacenamiento de archivos
from django.core.files.base import ContentFile  # Para convertir el archivo subido en contenido manejable
from django.utils.text import get_valid_filename
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.shortcuts import render
import uuid  # Para generar un identificador único
import uuid
from django.core.files.storage import FileSystemStorage  # Usamos FileSystemStorage para manejar archivos
from decimal import Decimal
from .validators import validar_precio, validar_cantidad
from decimal import Decimal  # Importar Decimal para manejar precios como decimales
from django.core.files.storage import FileSystemStorage
from .validators import validar_precio, validar_cantidad, validar_si_no
import os
from decimal import Decimal
from django.utils.text import get_valid_filename
from django.core.files.storage import FileSystemStorage
from django.core.files.base import ContentFile
from django.conf import settings
from django.contrib import messages
from .validators import validar_precio, validar_cantidad, validar_si_no
import openpyxl
import os
from decimal import Decimal
from django.utils.text import get_valid_filename
from django.core.files.storage import FileSystemStorage
from django.core.files.base import ContentFile
from django.conf import settings
from django.contrib import messages
from .validators import validar_precio, validar_cantidad, validar_si_no
import openpyxl
from django.contrib import messages
from django.shortcuts import render, redirect
from .forms import CargaMasivaForm
from .models import Producto
import openpyxl
from .validators import validar_precio, validar_cantidad
from django.conf import settings
import os
from django.core.files.storage import FileSystemStorage
from decimal import Decimal
import os
from decimal import Decimal
from django.utils.text import get_valid_filename
from django.core.files.storage import FileSystemStorage
from django.core.files.base import ContentFile
from django.conf import settings
from django.contrib import messages
from .validators import validar_precio, validar_cantidad, validar_rut
import openpyxl
import os
from decimal import Decimal
from django.utils.text import get_valid_filename
from django.core.files.storage import FileSystemStorage
from django.core.files.base import ContentFile
from django.conf import settings
from django.contrib import messages
from .validators import validar_precio, validar_cantidad, validar_rut
import openpyxl
import os
from decimal import Decimal
from django.utils.text import get_valid_filename
from django.core.files.storage import FileSystemStorage
from django.core.files.base import ContentFile
from django.conf import settings
from django.contrib import messages
from .validators import validar_precio, validar_cantidad, validar_rut
import openpyxl
import re
import os
from decimal import Decimal
from django.utils.text import get_valid_filename
from django.core.files.storage import FileSystemStorage
from django.core.files.base import ContentFile
from django.conf import settings
from django.contrib import messages
from .validators import validar_precio, validar_cantidad
import openpyxl
import re
import os





def carga_exitosa(request):
    return render(request, 'carga_masiva/carga_exitosa.html')


def validar_rut(nombre_archivo):
    """
    Valida si el nombre de un archivo corresponde a un RUT válido (formato chileno).
    El nombre debe ser algo como "12345678-9" o "12345678-K".
    """
    # Eliminar la extensión del archivo
    nombre_sin_extension = os.path.splitext(nombre_archivo)[0].strip()

    # Expresión regular para validar el formato del RUT (número seguido de guion y dígito verificador)
    rut_regex = r'^\d{1,8}-[\dkK]$'
    if not re.match(rut_regex, nombre_sin_extension):
        return False

    # Separar el número del dígito verificador
    rut_numero, rut_dv = nombre_sin_extension.split('-')

    # Validar el dígito verificador
    return validar_digito_verificador(rut_numero, rut_dv)

def validar_digito_verificador(rut_numero, rut_dv):
    """
    Valida el dígito verificador de un RUT chileno.
    """
    rut_numero = int(rut_numero)
    suma = 0
    multiplicador = 2

    while rut_numero > 0:
        suma += (rut_numero % 10) * multiplicador
        rut_numero //= 10
        multiplicador = 2 if multiplicador == 7 else multiplicador + 1

    resto = 11 - (suma % 11)
    if resto == 11:
        digito_calculado = '0'
    elif resto == 10:
        digito_calculado = 'K'
    else:
        digito_calculado = str(resto)

    return rut_dv.upper() == digito_calculado





def cargar_excel(request):
    if request.method == 'POST':
        form = CargaMasivaForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo_excel']
            nombre_archivo = archivo.name

            # Validar si el nombre del archivo corresponde a un RUT válido
            if not validar_rut(nombre_archivo):
                messages.error(request, 'El nombre del archivo no corresponde a un RUT válido. Nombra el archivo con el RUT de la institución, sin puntos y con guión')
                return render(request, 'carga_masiva/cargar_excel.html', {'form': form})

            try:
                # Guardar el archivo primero en la carpeta 'input'
                nombre_archivo_limpio = get_valid_filename(os.path.basename(archivo.name))
                fs = FileSystemStorage(location=settings.INPUT_FILES_DIR)
                
                # Verificar si el archivo ya existe y agregar un sufijo en caso de conflicto
                nombre_archivo_final = nombre_archivo_limpio
                contador = 1
                while fs.exists(nombre_archivo_final):
                    nombre_archivo_final = f"{os.path.splitext(nombre_archivo_limpio)[0]}_{contador}{os.path.splitext(nombre_archivo_limpio)[1]}"
                    contador += 1
                
                # Guardar el archivo en la carpeta 'input'
                file_path = fs.save(nombre_archivo_final, archivo)
                ruta_archivo = fs.path(file_path)  # Obtener la ruta completa donde se guardó el archivo

                # Procesar el archivo Excel cargado
                wb = openpyxl.load_workbook(ruta_archivo)
                hoja = wb.active

                # Definir las columnas obligatorias
                columnas_obligatorias = ["nombre", "descripcion", "precio", "cantidad"]

                # Obtener los encabezados (primera fila) y limpiar los nombres (eliminar espacios adicionales)
                encabezados = [str(cell.value).strip().lower() for cell in hoja[1]]  # Convertir a minúsculas y limpiar espacios

                # Validar si faltan las columnas obligatorias
                columnas_faltantes = [col for col in columnas_obligatorias if col not in encabezados]
                if columnas_faltantes:
                    messages.error(request, f'El archivo tiene columnas faltantes: {", ".join(columnas_faltantes)}.')
                    return render(request, 'carga_masiva/cargar_excel.html', {'form': form})

                # Identificar los índices de las columnas obligatorias
                idx_nombre = encabezados.index("nombre")
                idx_descripcion = encabezados.index("descripcion")
                idx_precio = encabezados.index("precio")
                idx_cantidad = encabezados.index("cantidad")

                # Continuar con la validación de filas y columnas
                errores = []  # Lista para almacenar los errores
                numero_fila = 2  # Comienza en 2 si la fila 1 es el encabezado

                for fila in hoja.iter_rows(min_row=2, values_only=True):
                    # Acceder solo a las columnas obligatorias usando los índices
                    nombre = fila[idx_nombre]
                    descripcion = fila[idx_descripcion]
                    precio = fila[idx_precio]
                    cantidad = fila[idx_cantidad]

                    # Validar la columna "precio" y "cantidad"
                    errores.extend(validar_precio(precio, numero_fila, "precio"))  # Validar precio
                    errores.extend(validar_cantidad(cantidad, numero_fila, "cantidad"))  # Validar cantidad

                    numero_fila += 1

                # Si hay errores, los mostramos al usuario
                if errores:
                    for error in errores:
                        messages.error(request, error)
                    return render(request, 'carga_masiva/cargar_excel.html', {'form': form})

                # Si no hay errores, guardar los datos en la base de datos
                numero_fila = 2  # Reiniciar para la creación
                for fila in hoja.iter_rows(min_row=2, values_only=True):
                    # Acceder solo a las columnas obligatorias usando los índices
                    nombre = fila[idx_nombre]
                    descripcion = fila[idx_descripcion]
                    precio = fila[idx_precio]
                    cantidad = fila[idx_cantidad]

                    Producto.objects.create(
                        nombre=nombre,
                        descripcion=descripcion,
                        precio=Decimal(str(precio).strip()),  # Convertir a Decimal antes de guardar
                        cantidad=int(cantidad)  # Convertir a int antes de guardar
                    )
                    numero_fila += 1

                messages.success(request, f'El archivo ha sido cargado exitosamente y guardado en {ruta_archivo}.')
                return redirect('carga_exitosa')

            except Exception as e:
                messages.error(request, f'Error al procesar el archivo: {e}')
        else:
            messages.error(request, 'El formulario no es válido.')

    else:
        form = CargaMasivaForm()

    return render(request, 'carga_masiva/cargar_excel.html', {'form': form})




